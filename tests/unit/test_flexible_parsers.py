"""Tests for flexible CSV/Excel column detection and the AWS/GCP CLI JSON parsers.

Real-world CSV/Excel/CLI-JSON exports never use the canonical
`type,name,id,provider` shape exactly — these tests cover the alias
detection, inference, and quirk-handling in `column_detection.py`, plus the
new `AWSCLIOutputParser` / `GCPCLIOutputParser`.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from migration_factory.domain.enums import CloudProvider
from migration_factory.parsers.extended import ExcelInventoryParser
from migration_factory.parsers.multi_format import AWSCLIOutputParser, CSVInventoryParser, GCPCLIOutputParser

FIXTURES = Path("tests/fixtures")


def _write_csv(tmp_path: Path, content: str, name: str = "inventory.csv") -> Path:
    path = tmp_path / name
    path.write_text(content, encoding="utf-8")
    return path


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


class TestFlexibleCSVParser:
    def test_standard_columns_still_work(self, tmp_path: Path) -> None:
        content = "type,name,id,provider,region\naws_instance,app-server,i-123,aws,us-east-1\n"
        path = _write_csv(tmp_path, content)
        result = CSVInventoryParser().parse(path)
        assert result.resource_count == 1
        r = result.resources[0]
        assert r.source_type == "aws_instance"
        assert r.name == "app-server"
        assert r.source_identifier == "i-123"
        assert r.source_provider == CloudProvider.AWS

    def test_aliased_columns_resource_type_name(self, tmp_path: Path) -> None:
        content = "resource_type,resource_name,resource_id\naws_vpc,main-vpc,vpc-999\n"
        path = _write_csv(tmp_path, content)
        result = CSVInventoryParser().parse(path)
        assert result.resource_count == 1
        r = result.resources[0]
        assert r.source_type == "aws_vpc"
        assert r.name == "main-vpc"
        assert r.source_identifier == "vpc-999"

    def test_aliased_columns_instanceid_instancetype(self, tmp_path: Path) -> None:
        # No 'type' column at all — must infer aws_instance from InstanceType's value.
        content = "InstanceId,InstanceType,Name,AvailabilityZone\ni-0abc123,t3.medium,web-1,us-east-1a\n"
        path = _write_csv(tmp_path, content)
        result = CSVInventoryParser().parse(path)
        assert result.resource_count == 1
        r = result.resources[0]
        assert r.source_type == "aws_instance"
        assert r.source_identifier == "i-0abc123"
        assert r.name == "web-1"
        assert r.source_provider == CloudProvider.AWS

    def test_aws_export_format_csv(self) -> None:
        result = CSVInventoryParser().parse(FIXTURES / "inventory_aws_export.csv")
        assert result.resource_count == 3
        types = {r.source_type for r in result.resources}
        assert types == {"aws_instance", "aws_s3_bucket", "aws_vpc"}
        assert all(r.source_provider == CloudProvider.AWS for r in result.resources)

    def test_infers_type_from_ec2_instance_type_column(self, tmp_path: Path) -> None:
        content = "Name,InstanceType,Region\nweb-1,m5.large,us-east-1\n"
        path = _write_csv(tmp_path, content)
        result = CSVInventoryParser().parse(path)
        assert result.resources[0].source_type == "aws_instance"

    def test_infers_provider_from_resource_type_prefix(self, tmp_path: Path) -> None:
        # No provider column at all; type is an explicit terraform-style google_ type.
        content = "type,name,id\ngoogle_compute_instance,vm-1,vm-001\n"
        path = _write_csv(tmp_path, content)
        result = CSVInventoryParser().parse(path)
        assert result.resources[0].source_provider == CloudProvider.GCP

    def test_handles_bom_in_csv(self, tmp_path: Path) -> None:
        content = "﻿type,name,id\naws_s3_bucket,my-bucket,my-bucket\n"
        path = tmp_path / "bom.csv"
        path.write_bytes(content.encode("utf-8"))
        result = CSVInventoryParser().parse(path)
        assert result.resource_count == 1
        assert result.resources[0].source_type == "aws_s3_bucket"

    def test_handles_empty_rows(self, tmp_path: Path) -> None:
        content = "type,name,id\naws_vpc,main,vpc-1\n,,\naws_subnet,sub,subnet-1\n"
        path = _write_csv(tmp_path, content)
        result = CSVInventoryParser().parse(path)
        assert result.resource_count == 2

    def test_collects_unknown_columns_as_attributes(self, tmp_path: Path) -> None:
        content = "type,name,id,Owner,CostCenter\naws_instance,app,i-1,platform-team,CC-100\n"
        path = _write_csv(tmp_path, content)
        result = CSVInventoryParser().parse(path)
        attrs = result.resources[0].attributes
        assert attrs.get("Owner") == "platform-team"
        assert attrs.get("CostCenter") == "CC-100"


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def _write_xlsx(tmp_path: Path, rows: list[list[object]], name: str = "inventory.xlsx") -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    path = tmp_path / name
    wb.save(str(path))
    return path


class TestFlexibleExcelParser:
    def test_standard_columns_still_work(self, tmp_path: Path) -> None:
        path = _write_xlsx(tmp_path, [
            ["type", "name", "id", "provider", "region"],
            ["aws_instance", "app-server", "i-123", "aws", "us-east-1"],
            ["aws_vpc", "main", "vpc-456", "aws", "us-east-1"],
        ])
        result = ExcelInventoryParser().parse(path)
        assert result.resource_count == 2
        assert result.resources[0].source_type == "aws_instance"

    def test_title_row_detection(self, tmp_path: Path) -> None:
        path = _write_xlsx(tmp_path, [
            ["AWS Resource Inventory", None, None, None, None],
            ["InstanceId", "InstanceType", "AvailabilityZone", "State", "Name"],
            ["i-0aaa111", "t3.medium", "us-east-1a", "running", "web-1"],
            ["i-0bbb222", "m5.large", "us-east-1b", "running", "app-1"],
            ["i-0ccc333", "t3.small", "us-east-1b", "running", "worker-1"],
        ])
        result = ExcelInventoryParser().parse(path)
        assert result.resource_count == 3
        assert result.resources[0].source_type == "aws_instance"
        assert result.resources[0].name == "web-1"
        assert result.resources[0].source_identifier == "i-0aaa111"

    def test_aliased_column_names(self, tmp_path: Path) -> None:
        path = _write_xlsx(tmp_path, [
            ["Asset Type", "Asset Name", "Asset Id"],
            ["aws_lambda_function", "process-orders", "arn:aws:lambda:us-east-1:123:function:process-orders"],
        ])
        result = ExcelInventoryParser().parse(path)
        assert result.resource_count == 1
        r = result.resources[0]
        assert r.source_type == "aws_lambda_function"
        assert r.name == "process-orders"

    def test_numeric_cell_values_converted(self, tmp_path: Path) -> None:
        path = _write_xlsx(tmp_path, [
            ["type", "name", "id", "port"],
            ["aws_instance", "app", "i-1", 8080],
        ])
        result = ExcelInventoryParser().parse(path)
        assert result.resources[0].attributes.get("port") == "8080"

    def test_empty_cells_handled(self, tmp_path: Path) -> None:
        path = _write_xlsx(tmp_path, [
            ["type", "name", "id", "provider"],
            ["aws_instance", "app", "i-1", None],
        ])
        result = ExcelInventoryParser().parse(path)
        assert result.resource_count == 1
        # No provider column value and no region to infer from -> UNKNOWN, not a crash.
        assert result.resources[0].source_provider in {CloudProvider.UNKNOWN, CloudProvider.AWS}

    def test_uses_first_sheet_with_data(self, tmp_path: Path) -> None:
        import openpyxl

        wb = openpyxl.Workbook()
        empty_ws = wb.active
        empty_ws.title = "Empty"
        data_ws = wb.create_sheet("Data")
        data_ws.append(["type", "name", "id"])
        data_ws.append(["aws_vpc", "main", "vpc-1"])
        path = tmp_path / "multi_sheet.xlsx"
        wb.save(str(path))

        result = ExcelInventoryParser().parse(path)
        assert result.resource_count == 1
        assert result.resources[0].source_type == "aws_vpc"


@pytest.fixture
def inventory_flexible_xlsx(tmp_path: Path) -> Path:
    """tests/fixtures/inventory_flexible.xlsx, generated at test time (per
    spec) rather than checked in as a binary: a title row, headers in row 2,
    and 3 rows of EC2-instance-shaped data.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["AWS Resource Inventory"])
    ws.append(["InstanceId", "InstanceType", "AvailabilityZone", "State", "Name"])
    ws.append(["i-0flex111aaa", "t3.medium", "us-east-1a", "running", "flex-web-1"])
    ws.append(["i-0flex222bbb", "m5.large", "us-east-1b", "running", "flex-app-1"])
    ws.append(["i-0flex333ccc", "t3.small", "us-east-1b", "running", "flex-worker-1"])
    path = tmp_path / "inventory_flexible.xlsx"
    wb.save(str(path))
    return path


def test_inventory_flexible_xlsx_fixture(inventory_flexible_xlsx: Path) -> None:
    result = ExcelInventoryParser().parse(inventory_flexible_xlsx)
    assert result.resource_count == 3
    assert all(r.source_type == "aws_instance" for r in result.resources)
    assert {r.name for r in result.resources} == {"flex-web-1", "flex-app-1", "flex-worker-1"}


# ---------------------------------------------------------------------------
# AWS CLI JSON
# ---------------------------------------------------------------------------


class TestAWSCLIOutputParser:
    def test_supports_describe_instances_output(self) -> None:
        assert AWSCLIOutputParser().supports(FIXTURES / "aws_cli_instances.json") is True

    def test_parses_reservations_to_aws_instance(self) -> None:
        result = AWSCLIOutputParser().parse(FIXTURES / "aws_cli_instances.json")
        assert result.resource_count == 3
        assert all(r.source_type == "aws_instance" for r in result.resources)
        assert all(r.source_provider == CloudProvider.AWS for r in result.resources)

    def test_extracts_name_from_tags(self) -> None:
        result = AWSCLIOutputParser().parse(FIXTURES / "aws_cli_instances.json")
        names = {r.name for r in result.resources}
        assert names == {"web-server-1", "app-server-1", "worker-1"}

    def test_parses_vpcs(self, tmp_path: Path) -> None:
        data = {"Vpcs": [{"VpcId": "vpc-1", "CidrBlock": "10.0.0.0/16", "IsDefault": False,
                          "Tags": [{"Key": "Name", "Value": "my-vpc"}]}]}
        path = tmp_path / "vpcs.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        result = AWSCLIOutputParser().parse(path)
        assert result.resource_count == 1
        r = result.resources[0]
        assert r.source_type == "aws_vpc"
        assert r.name == "my-vpc"
        assert r.attributes["cidr_block"] == "10.0.0.0/16"

    def test_parses_subnets(self, tmp_path: Path) -> None:
        data = {"Subnets": [{"SubnetId": "subnet-1", "VpcId": "vpc-1", "CidrBlock": "10.0.1.0/24",
                             "AvailabilityZone": "us-east-1a"}]}
        path = tmp_path / "subnets.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        result = AWSCLIOutputParser().parse(path)
        assert result.resource_count == 1
        r = result.resources[0]
        assert r.source_type == "aws_subnet"
        assert r.attributes["az"] == "us-east-1a"

    def test_parses_security_groups(self, tmp_path: Path) -> None:
        data = {"SecurityGroups": [{"GroupId": "sg-1", "GroupName": "web-sg", "Description": "web",
                                     "VpcId": "vpc-1", "IpPermissions": [], "IpPermissionsEgress": []}]}
        path = tmp_path / "sgs.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        result = AWSCLIOutputParser().parse(path)
        assert result.resource_count == 1
        r = result.resources[0]
        assert r.source_type == "aws_security_group"
        assert r.name == "web-sg"

    def test_parses_s3_buckets(self, tmp_path: Path) -> None:
        data = {"Buckets": [{"Name": "my-bucket", "CreationDate": "2024-01-01T00:00:00Z"}]}
        path = tmp_path / "buckets.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        result = AWSCLIOutputParser().parse(path)
        assert result.resource_count == 1
        r = result.resources[0]
        assert r.source_type == "aws_s3_bucket"
        assert r.source_identifier == "my-bucket"

    def test_parses_rds_instances(self, tmp_path: Path) -> None:
        data = {"DBInstances": [{"DBInstanceIdentifier": "orders-db", "Engine": "postgres",
                                  "DBInstanceClass": "db.t3.medium", "AllocatedStorage": 100, "MultiAZ": True}]}
        path = tmp_path / "rds.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        result = AWSCLIOutputParser().parse(path)
        assert result.resource_count == 1
        r = result.resources[0]
        assert r.source_type == "aws_db_instance"
        assert r.attributes["engine"] == "postgres"

    def test_parses_combined_output(self) -> None:
        result = AWSCLIOutputParser().parse(FIXTURES / "aws_cli_combined.json")
        assert result.resource_count == 7
        types = [r.source_type for r in result.resources]
        assert types.count("aws_instance") == 3
        assert types.count("aws_vpc") == 1
        assert types.count("aws_subnet") == 2
        assert types.count("aws_security_group") == 1

    def test_parses_query_output(self) -> None:
        result = AWSCLIOutputParser().parse(FIXTURES / "aws_cli_query_output.json")
        assert result.resource_count == 2
        assert all(r.source_type == "aws_instance" for r in result.resources)

    def test_all_3_fixture_files(self) -> None:
        parser = AWSCLIOutputParser()
        for fixture, expected_count in (
            ("aws_cli_instances.json", 3),
            ("aws_cli_combined.json", 7),
            ("aws_cli_query_output.json", 2),
        ):
            path = FIXTURES / fixture
            assert parser.supports(path) is True, fixture
            result = parser.parse(path)
            assert result.resource_count == expected_count, fixture


# ---------------------------------------------------------------------------
# GCP CLI JSON
# ---------------------------------------------------------------------------


class TestGCPCLIOutputParser:
    def test_parses_gcloud_compute_instances(self, tmp_path: Path) -> None:
        data = [
            {
                "name": "vm-1",
                "machineType": "https://www.googleapis.com/compute/v1/projects/p/zones/us-central1-a/machineTypes/e2-medium",
                "zone": "https://www.googleapis.com/compute/v1/projects/p/zones/us-central1-a",
                "status": "RUNNING",
                "labels": {"env": "prod"},
            }
        ]
        path = tmp_path / "instances.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        assert GCPCLIOutputParser().supports(path) is True
        result = GCPCLIOutputParser().parse(path)
        assert result.resource_count == 1
        r = result.resources[0]
        assert r.source_type == "google_compute_instance"
        assert r.source_provider == CloudProvider.GCP
        assert r.attributes["machine_type"] == "e2-medium"
        assert r.attributes["zone"] == "us-central1-a"

    def test_parses_gcloud_storage_buckets(self, tmp_path: Path) -> None:
        data = [{"name": "my-gcs-bucket", "location": "US", "storageClass": "STANDARD", "labels": {}}]
        path = tmp_path / "buckets.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        assert GCPCLIOutputParser().supports(path) is True
        result = GCPCLIOutputParser().parse(path)
        assert result.resource_count == 1
        r = result.resources[0]
        assert r.source_type == "google_storage_bucket"
        assert r.attributes["storage_class"] == "STANDARD"

    def test_infers_resource_type_from_gcloud_output(self, tmp_path: Path) -> None:
        sql_data = [{"name": "orders-sql", "databaseVersion": "POSTGRES_14", "region": "us-central1",
                     "settings": {"tier": "db-custom-2-8192"}}]
        network_data = [{"name": "default", "autoCreateSubnetworks": True, "subnetworks": []}]

        sql_path = tmp_path / "sql.json"
        sql_path.write_text(json.dumps(sql_data), encoding="utf-8")
        network_path = tmp_path / "network.json"
        network_path.write_text(json.dumps(network_data), encoding="utf-8")

        sql_result = GCPCLIOutputParser().parse(sql_path)
        assert sql_result.resources[0].source_type == "google_sql_database_instance"

        network_result = GCPCLIOutputParser().parse(network_path)
        assert network_result.resources[0].source_type == "google_compute_network"


# ---------------------------------------------------------------------------
# End to end
# ---------------------------------------------------------------------------


class TestEndToEnd:
    def test_full_pipeline_aws_cli_json_to_gcp_terraform(self, tmp_path: Path) -> None:
        from migration_factory.assessment.engine import AssessmentEngine
        from migration_factory.core.config import Settings
        from migration_factory.pipeline import IngestionPipeline
        from migration_factory.terraform_gen.engine import TerraformGenerator
        from migration_factory.translation.engine import TranslationEngine
        from migration_factory.translation.matrix import load_builtin_matrix

        ingestion = IngestionPipeline(settings=Settings()).run(FIXTURES / "aws_cli_instances.json")
        assert len(ingestion.graph.resources) == 3

        matrix = load_builtin_matrix(CloudProvider.AWS, CloudProvider.GCP)
        translation = TranslationEngine(matrix=matrix).translate(ingestion.graph)
        AssessmentEngine().assess(ingestion.graph, translation)

        terraform = TerraformGenerator(target_provider=CloudProvider.GCP, project_id="test-project").generate(
            ingestion.graph, translation
        )
        combined = "\n".join(f.content for f in terraform.files)
        assert "google_compute_instance" in combined

    def test_full_pipeline_flexible_csv_to_gcp_terraform(self) -> None:
        from migration_factory.core.config import Settings
        from migration_factory.pipeline import IngestionPipeline

        ingestion = IngestionPipeline(settings=Settings()).run(FIXTURES / "inventory_flexible.csv")
        assert len(ingestion.graph.resources) >= 3
        types = {r.source_type for r in ingestion.graph.resources.values()}
        assert "aws_instance" in types or "aws_s3_bucket" in types
